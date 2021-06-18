import openpyxl
import os.path
from backend.core import alg
from backend.core.company import Company
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill

# Fetch the next alphabetical symbol in 'coc.xlsx'
def get_next_symbol():
    prev_symbol = ""
    for row in ws.iter_rows(ws.max_row, ws.max_row):
        prev_symbol = row[0].value
    symbol = ""
    f = open("backend/tickers.txt", "r")
    lines = f.readlines()
    print("prev_symbol: " + str(prev_symbol))
    for i in range(0, len(lines)):
        items = lines[i].split("|")
        if items[1] == prev_symbol:
            # Get next as long as it's not OOB
            if (i + 1) < len(lines):
                symbol = lines[i + 1].split("|")[1]
            break
    f.close()
    return symbol


# Color codes criteria columns with green/red for if they pass/fail
def color_code_row(row_num, ws, colors):
    green = PatternFill(fill_type="solid", start_color="3CB371", end_color="3CB371")
    red = PatternFill(fill_type="solid", start_color="CD5C5C", end_color="CD5C5C")
    white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    yellow = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00") 
    idx = 0
    for alpha in range(ord("A"), ord("N") + 1):
        curr_cell = ws[chr(alpha) + str(row_num)]
        if "green" in colors[idx]:
            curr_cell.fill = green
        elif "yellow" in colors[idx]:
            curr_cell.fill = yellow
        elif "red" in colors[idx]:
            curr_cell.fill = red
        elif "white" in colors[idx]:
            curr_cell.fill = white
        else:
            curr_cell.fill = white
        idx += 1


# Creates an array of the colors for the columns of a row
def generate_cell_colors(graham_ratio, bvps_ratio, health_result):
    colors = ["white", "white", "white", "white"]  # symbol, score, sector, price
    if graham_ratio >= 1:
        colors.append("green")
    else:
        colors.append("red")
    if bvps_ratio >= 1:
        colors.append("green")
    else:
        colors.append("red")
    colors.append("white")  # dividend yield
    for criteria in health_result:
        if "green" in criteria:
            colors.append("green")
        elif "yellow" in criteria:
            colors.append("yellow")
        elif "red" in criteria:
            colors.append("red")
        else:
            colors.append("white")
    return colors


# Adds a new row for this symbol to the end of the excel file
def update(symbol, company):
    dest_filename = "coc.xlsx"
    overwrite_row = None
    if os.path.isfile(dest_filename):
        wb = load_workbook(filename=dest_filename)
        ws = wb.active
        for idx in range(2, ws.max_row + 1):
            curr_symbol = ws["A" + str(idx)].value
            if curr_symbol.lower() == symbol.lower():
                overwrite_row = idx
                break
    else:
        wb = Workbook()
        ws = wb.active
        # Initializes the sheet name and header row
        ws.title = "Analysis"
        ws.append(
            [
                "",
                "Score",
                "Sector",
                "Price",
                "Graham Num (vs. price)",
                "BVPS (vs. price)",
                "Div. Yield (Payout Ratio)",
                "Criteria 1 (Market Cap vs. Net Asset Value)",
                "Criteria 2 (Annual EPS Growth)",
                "Criteria 3 (Earnings Deficits)",
                "Criteria 4 (Current Ratio)",
                "Criteria 5 (Cheapness of Earnings)",
                "Criteria 6 (Strength of Sales)",
                "Criteria 7 (Dividend Reliability)",
            ]
        )
        # Resize column widths to show full column titles. Skips (empty) Col A.
        for alpha in range(ord("B"), ord("M") + 1):
            curr_char = chr(alpha)
            title_width = len(ws[curr_char + "1"].value)
            if title_width > 12:
                ws.column_dimensions[curr_char].width = title_width
            else:
                ws.column_dimensions[curr_char].width = 12
        # Freezes the top row of the excel file
        wb["Analysis"].freeze_panes = "A2"
    health_result = company.health_check()
    # Ratios relative to price
    graham_ratio = bvps_ratio = 0.00
    if company.graham_num is not None and company.price is not None:
        graham_ratio = round(company.graham_num / company.price, 2)
    if company.bvps is not None and company.price is not None:
        bvps_ratio = round(company.bvps / company.price, 2)
    colors = generate_cell_colors(graham_ratio, bvps_ratio, health_result)
    for i in range(0, len(health_result)):
        health_result[i] = (
            health_result[i]
            .replace("[green]", "")
            .replace("[/green]", "")
            .replace("[red]", "")
            .replace("[/red]", "")
            .replace("[yellow]", "")
            .replace("[/yellow]", "")
        )
        health_result[i] = health_result[i][3:]  # Removes the 'CX: ' prefix
    new_row = [
        company.symbol.upper(),
        str(company.score),
        company.sector,
        str(company.price),
        str(company.graham_num) + " (" + str(graham_ratio) + ")",
        str(company.bvps) + " (" + str(bvps_ratio) + ")",
        str(company.div_yield) + " (" + str(company.payout_ratio) + ")",
        health_result[0],
        health_result[1],
        health_result[2],
        health_result[3],
        health_result[4],
        health_result[5],
        health_result[6],
    ]
    # Either overwrites the row for the symbol or adds a new row for it
    idx = 0
    if overwrite_row != None:
        for col, val in enumerate(new_row, start=1):
            ws.cell(row=overwrite_row, column=col).value = val
            color_code_row(overwrite_row, ws, colors)
    else:
        ws.append(new_row)
        color_code_row(ws.max_row, ws, colors)
    wb.save(filename=dest_filename)
    return
