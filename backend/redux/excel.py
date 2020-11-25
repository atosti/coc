import pandas as pd
import openpyxl
import os.path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill

# Fetch the next alphabetical symbol in 'coc.xlsx'
def getNextSymbol():
    prevSymbol = ''
    for row in ws.iter_rows(ws.max_row, ws.max_row):
        prevSymbol = row[0].value
    symbol = ''
    f = open("backend/tickers.txt", "r")
    lines = f.readlines()
    print("Prevsymbol: " + str(prevSymbol))
    for i in range (0, len(lines)):
        items = lines[i].split('|')
        if items[1] == prevSymbol:
            # Get next as long as it's not OOB
            if (i + 1) < len(lines):
                symbol = lines[i+1].split('|')[1]
            break;
    f.close()
    return symbol

# Adds a new row for this symbol to the end of the excel file
def update(symbol, dataDict):
    dest_filename = 'coc.xlsx'
    overwriteRow = None
    if os.path.isfile(dest_filename):
        wb = load_workbook(filename = dest_filename)
        ws = wb.active
        for idx in range(2, ws.max_row + 1):
            currSymbol = ws['A' + str(idx)].value
            if currSymbol.lower() == symbol.lower():
                overwriteRow = idx
                break;
    else:
        wb = Workbook()
        ws = wb.active
        # Initializes the sheet name and header row
        ws.title = 'Analysis'
        ws.append(['', 'Score', 'Sector', 'Graham Num.', 'Price', 'Div. Yield',
            'Sales > $700M', 'Curr. Ratio >= 2', 'No Missed Dividend(5yrs)', 
            'No Earnings Deficit(5yrs)', 'EPS avg > 33%(5yrs)', 'Cheap Assets', 
            'P/E < 15'])
        # Resize column widths to show full column titles. Skips (empty) Col A.
        for alpha in range(ord('B'), ord('M') + 1):
            currChar = chr(alpha)
            titleWidth = len(ws[currChar + '1'].value)
            if titleWidth > 12:
                ws.column_dimensions[currChar].width = titleWidth
            else:
                ws.column_dimensions[currChar].width = 12
        # Freezes the top row of the excel file
        wb['Analysis'].freeze_panes = 'A2'
    # Stores the row being added
    newRow = [dataDict['symbol'].upper(), dataDict['score'],
        dataDict['sector'], dataDict['grahamNum'], dataDict['price'], 
        dataDict['divYield'], dataDict['goodSales'], dataDict['goodCurrRatio'], 
        dataDict['goodDividend'], dataDict['goodEps'], 
        dataDict['goodEpsGrowth'], dataDict['goodAssets'], 
        dataDict['goodPeRatio']
    ]
    # Either overwrites the row for the symbol or adds a new row for it
    if overwriteRow != None:
        for col, val in enumerate(newRow, start=1):
            ws.cell(row=overwriteRow, column=col).value = val
    else:
        ws.append(newRow)
    # Fills True/False cells Green/Red in new row for readability
    greenFill = PatternFill(fill_type='solid', start_color='3CB371', 
    end_color='3CB371')
    redFill = PatternFill(fill_type='solid', start_color='CD5C5C', 
        end_color='CD5C5C')
    for alpha in range(ord('G'), ord('M') + 1):
        currCell = ws[chr(alpha) + str(ws.max_row)]
        if currCell.value:
            currCell.fill = greenFill
        else:
            currCell.fill = redFill
    wb.save(filename = dest_filename)
    return